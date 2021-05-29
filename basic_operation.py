import openpyxl
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter

# Instantiate worksheet
wb = load_workbook("Grade.xlsx") # load the excel workbook
ws = wb.active  # create worksheet

# Get values from excel
print(ws['A2'].value)

# Change a value from excel
ws['A2'] = 'Paul'
wb.save("Grade.xlsx")

# Access different sheet
print(wb.sheetnames)
print(wb['Sheet2'])
ws = wb["Sheet2"]
print(ws["A1"].value)

# Create a worksheet and write something on it
wb = Workbook()
ws = wb.active
ws.title = "Data"  # add sheet title

ws.append(["Paul", "Is", "Great", "!"])
ws.append(["Paul", "Is", "Great", "!"])
ws.append(["Paul", "Is", "Great", "!"])
ws.append(["Paul", "Is", "Great", "!"])
ws.append(["End"])

wb.save("Paul.xlsx")

# Write excel in a more efficient way
wb = load_workbook('Paul.xlsx')
ws = wb.active

for row in range(1,11):
    # # To access each cell like "A1", "B2",
    # # either use python built-in chr() func like so:
    # for col in range(0,4):
    #     char= chr(65 + col)  # chr(65) --> A , chr(97) --> a

    # or use get_column_letter()
    for col in range(1,5):
        char = get_column_letter(col)
        ws[char + str(row)] = char + str(row)  # get_column_letter[1] --> A
wb.save('Paul.xlsx')


wb = load_workbook('Paul.xlsx')
ws = wb.active

# Merge cells
ws.merge_cells("A1:D1")
ws.unmerge_cells("A1:D1")

# Insert column
ws.insert_cols(2) # insert a column at col 2

# Insert row
ws.insert_rows(2) # insert a row at row 2

# delete col
ws.delete_cols(2) # delete a column at col 2

# delete row
ws.delete_rows(2) # delete a row at row 2

# Copy and move cells
ws.move_range("C1:D11", rows=2, cols=2)

wb.save('Paul.xlsx')
